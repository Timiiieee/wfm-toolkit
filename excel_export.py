"""Build an interactive Erlang-C staffing workbook (output/staffing_model.xlsx).

The workbook implements the full Erlang-C model as LIVE Excel formulas, so a
non-technical user can open it, change the input cells (call volume, AHT,
target service level, max occupancy, shrinkage), and watch the recommended
agent count, service level, occupancy, ASA, scheduled headcount, and the charts
recalculate instantly. No Python needed to use it.

Layout:
  - Inputs block (editable, highlighted): demand + service target + the two
    real-world planning levers (max-occupancy cap, shrinkage)
  - Derived values (interval seconds, offered load in Erlangs)
  - Result block (recommended on-phone agents, resulting SL/occupancy/ASA, and
    scheduled headcount after shrinkage)
  - Staffing table: Erlang-B (iterative), occupancy, Erlang-C, SL, ASA per agent
  - Two line charts: service level vs agents, occupancy vs agents

Usage:
    python3 excel_export.py [output_path]
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.properties import CalcProperties

MAX_AGENTS = 100  # staffing table covers n = 1..MAX_AGENTS

# Default inputs (a busy interval, so the model shows a meaningful result).
DEFAULT_CALLS = 240
DEFAULT_AHT = 240
DEFAULT_INTERVAL_MIN = 30
DEFAULT_TARGET_SL = 0.80
DEFAULT_TARGET_SEC = 20
DEFAULT_MAX_OCCUPANCY = 0.85   # standard WFM occupancy ceiling
DEFAULT_SHRINKAGE = 0.30       # standard planning shrinkage

# Styles
TITLE_FONT = Font(size=14, bold=True)
HEAD_FONT = Font(bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # soft yellow = editable
RESULT_FILL = PatternFill("solid", fgColor="D9EAD3")  # soft green = output
HEAD_FILL = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build(path: str) -> None:
    wb = Workbook()
    # Force Excel/LibreOffice to recalculate all formulas when the file opens,
    # so charts populate without the user pressing anything.
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

    ws = wb.active
    ws.title = "Staffing Model"

    ws["A1"] = "Erlang-C Call Center Staffing Model"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Edit the yellow input cells. Everything else updates automatically."
    ws["A2"].font = Font(italic=True, color="666666")

    # --- Inputs ---
    ws["A4"] = "INPUTS"
    ws["A4"].font = HEAD_FONT
    inputs = [
        ("Calls per interval", DEFAULT_CALLS, "0"),
        ("Average handle time (sec)", DEFAULT_AHT, "0"),
        ("Interval length (min)", DEFAULT_INTERVAL_MIN, "0"),
        ("Target service level", DEFAULT_TARGET_SL, "0%"),
        ("Answer within (sec)", DEFAULT_TARGET_SEC, "0"),
        ("Max occupancy (1 = no cap)", DEFAULT_MAX_OCCUPANCY, "0%"),
        ("Shrinkage", DEFAULT_SHRINKAGE, "0%"),
    ]
    row = 5
    for label, value, fmt in inputs:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=value)
        c.fill = INPUT_FILL
        c.border = BORDER
        c.number_format = fmt
        row += 1
    # Named references for readability
    CALLS, AHT, INTERVAL_MIN = "$B$5", "$B$6", "$B$7"
    TARGET_SL, TARGET_SEC = "$B$8", "$B$9"
    MAXOCC, SHRINK = "$B$10", "$B$11"
    # Safe forms: max occupancy collapses to 1 (no cap) if blank/invalid;
    # shrinkage collapses to 0 if blank/invalid.
    MAXOCC_SAFE = f"IF(OR({MAXOCC}<=0,{MAXOCC}>1),1,{MAXOCC})"
    SHRINK_SAFE = f"IF(OR({SHRINK}<0,{SHRINK}>=1),0,{SHRINK})"

    # --- Derived ---
    ws["A13"] = "DERIVED"
    ws["A13"].font = HEAD_FONT
    ws["A14"] = "Interval length (sec)"
    ws["B14"] = f"={INTERVAL_MIN}*60"
    ws["A15"] = "Offered load (Erlangs)"
    ws["B15"] = f"={CALLS}*{AHT}/B14"
    ws["B15"].number_format = "0.00"
    LOAD = "$B$15"

    # --- Staffing table ---
    tbl_head = 17
    headers = ["Agents (n)", "Erlang-B", "Occupancy", "Erlang-C P(wait)",
               "Service level", "ASA (sec)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=tbl_head, column=j, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    first_data = tbl_head + 1            # seed row (n=0)
    ws.cell(row=first_data, column=1, value=0)
    ws.cell(row=first_data, column=2, value=1.0)  # Erlang-B B(0)=1

    n_first = first_data + 1             # n = 1
    n_last = first_data + MAX_AGENTS     # n = MAX_AGENTS
    for i in range(MAX_AGENTS):
        r = n_first + i
        prev = r - 1
        n_cell = f"A{r}"
        b_prev = f"B{prev}"
        ws.cell(row=r, column=1, value=i + 1)  # agents n
        # Erlang-B iterative: B(n) = (A*Bprev)/(n + A*Bprev)
        ws.cell(row=r, column=2,
                value=f"=({LOAD}*{b_prev})/({n_cell}+{LOAD}*{b_prev})")
        # Occupancy rho = A/n
        ws.cell(row=r, column=3, value=f"={LOAD}/{n_cell}")
        # Erlang-C: if rho>=1 -> 1 else B/(1 - rho*(1-B))
        ws.cell(row=r, column=4,
                value=f"=IF(C{r}>=1,1,B{r}/(1-C{r}*(1-B{r})))")
        # Service level: guards for zero load and understaffing, clamped 0..1
        ws.cell(row=r, column=5,
                value=(f"=IF({LOAD}=0,1,IF(C{r}>=1,0,"
                       f"MAX(0,MIN(1,1-D{r}*EXP(-({n_cell}-{LOAD})*({TARGET_SEC}/{AHT}))))))"))
        # ASA seconds: blank-ish when understaffed
        ws.cell(row=r, column=6,
                value=f'=IF(C{r}>=1,"n/a",D{r}*{AHT}/({n_cell}-{LOAD}))')
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=3).number_format = "0%"
        ws.cell(row=r, column=5).number_format = "0.0%"
        ws.cell(row=r, column=4).number_format = "0.000"
        ws.cell(row=r, column=2).number_format = "0.0000"
        ws.cell(row=r, column=6).number_format = "0.0"

    # --- Result block (placed to the right of the inputs) ---
    n_range = f"A{n_first}:A{n_last}"
    sl_range = f"E{n_first}:E{n_last}"
    occ_range = f"C{n_first}:C{n_last}"
    asa_range = f"F{n_first}:F{n_last}"

    ws["D4"] = "RESULT"
    ws["D4"].font = HEAD_FONT
    # Minimum agents meeting the service-level target. MINIFS returns 0 if no
    # row in the table clears the target.
    minifs = f"MINIFS({n_range},{sl_range},\">=\"&{TARGET_SL})"
    # Floor implied by the occupancy cap: n must be >= load / max_occupancy.
    occ_floor = f"CEILING({LOAD}/({MAXOCC_SAFE}),1)"
    # Recommended on-phone agents = the larger of the SL-driven count and the
    # occupancy-cap floor. Readable message instead of #N/A if infeasible.
    recommended = (f"=IF({minifs}=0,\"raise agent cap / lower target\","
                   f"MAX({minifs},{occ_floor}))")
    results = [
        ("Recommended agents (on phone)", recommended, "0"),
        ("Service level achieved",
         f"=IFERROR(INDEX({sl_range},MATCH($E$5,{n_range},0)),\"n/a\")", "0.0%"),
        ("Occupancy at that staffing",
         f"=IFERROR(INDEX({occ_range},MATCH($E$5,{n_range},0)),\"n/a\")", "0%"),
        ("ASA at that staffing (sec)",
         f"=IFERROR(INDEX({asa_range},MATCH($E$5,{n_range},0)),\"n/a\")", "0.0"),
        ("Scheduled headcount (after shrinkage)",
         f"=IF(ISNUMBER($E$5),$E$5/(1-({SHRINK_SAFE})),\"n/a\")", "0.0"),
    ]
    rr = 5
    for label, formula, fmt in results:
        ws.cell(row=rr, column=4, value=label)
        c = ws.cell(row=rr, column=5, value=formula)
        c.fill = RESULT_FILL
        c.border = BORDER
        c.number_format = fmt
        rr += 1

    # --- Charts ---
    # Categories include the seed row (n=0) so the category range is the same
    # length as the data series (which begins at the header row via
    # titles_from_data). This keeps each plotted point aligned with its agent
    # count instead of shifting by one position.
    cats = Reference(ws, min_col=1, min_row=first_data, max_row=n_last)

    sl_chart = LineChart()
    sl_chart.title = "Service level vs agents"
    sl_chart.x_axis.title = "Agents"
    sl_chart.y_axis.title = "Service level"
    sl_chart.height = 8
    sl_chart.width = 15
    sl_data = Reference(ws, min_col=5, min_row=tbl_head, max_row=n_last)
    sl_chart.add_data(sl_data, titles_from_data=True)
    sl_chart.set_categories(cats)
    ws.add_chart(sl_chart, "H4")

    occ_chart = LineChart()
    occ_chart.title = "Occupancy vs agents"
    occ_chart.x_axis.title = "Agents"
    occ_chart.y_axis.title = "Occupancy"
    occ_chart.height = 8
    occ_chart.width = 15
    occ_data = Reference(ws, min_col=3, min_row=tbl_head, max_row=n_last)
    occ_chart.add_data(occ_data, titles_from_data=True)
    occ_chart.set_categories(cats)
    ws.add_chart(occ_chart, "H20")

    # Column widths
    widths = {"A": 26, "B": 12, "C": 12, "D": 30, "E": 16, "F": 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    print(f"Wrote interactive workbook to {path}")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "output", "staffing_model.xlsx"
    )
    build(out)
